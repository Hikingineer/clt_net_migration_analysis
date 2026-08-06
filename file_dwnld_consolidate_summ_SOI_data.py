import os
import requests
import pandas as pd
from openpyxl import Workbook

# ---- config ----
states = ["nc"]

overall_start_yy = 11
overall_end_yy = 22  # so last filename 2223nc.xlsx

IN_SHEET = "County Inflow"
OUT_SHEET = "County Outflow"

URL_BASE = "https://www.irs.gov/pub/irs-soi/"
DOWNLOAD_DIR = "downloads"
os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def download_excel(url, out_path):
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with open(out_path, "wb") as f:
        f.write(r.content)


def ext_for_startyy(startyy: int) -> str:
    # xls till 19 (startyy <= 19); xlsx from 20 (startyy >= 20)
    return ".xlsx" if startyy >= 20 else ".xls"


def url_for_year_state(startyy: int, state_code: str):
    endyy = startyy + 1
    ext = ext_for_startyy(startyy)
    filename = f"{startyy}{endyy}{state_code}{ext}"
    return URL_BASE + filename, filename


def read_sheet(path, sheet_name):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return pd.read_excel(path, sheet_name=sheet_name, engine="xlrd")
    return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")


def to_float(x):
    try:
        if x is None:
            return 0.0
        if isinstance(x, str) and x.strip() == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0


def to_int_like(x):
    try:
        if x is None:
            return None
        return int(float(x))
    except Exception:
        return None


def days_in_year(y: int) -> int:
    return 366 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 365


for state in states:
    out_xlsx = f"{state}_county_inflow_outflow.xlsx"

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "raw data"

    wrote_headers = False
    next_row = 2  # leave row 1 for headers

    # ---- build "raw data" ----
    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy

        url, filename = url_for_year_state(startyy, state)
        local_path = os.path.join(DOWNLOAD_DIR, filename)

        print("Downloading:", url)
        download_excel(url, local_path)

        inflow = read_sheet(local_path, IN_SHEET)
        outflow = read_sheet(local_path, OUT_SHEET)

        if not wrote_headers:
            ws_raw.cell(row=1, column=1, value="Year")
            for j, colname in enumerate(inflow.columns.tolist(), start=2):
                ws_raw.cell(row=1, column=j, value=colname)
            wrote_headers = True
            next_row = 2

        # inflow
        for _, r in inflow.iterrows():
            ws_raw.cell(row=next_row, column=1, value=year_value)
            for j, val in enumerate(r.tolist(), start=2):
                ws_raw.cell(row=next_row, column=j, value=val)
            next_row += 1

        # outflow
        for _, r in outflow.iterrows():
            ws_raw.cell(row=next_row, column=1, value=year_value)
            for j, val in enumerate(r.tolist(), start=2):
                ws_raw.cell(row=next_row, column=j, value=val)
            next_row += 1

    raw_last_row = ws_raw.max_row
    if raw_last_row < 2:
        wb.save(out_xlsx)
        raise RuntimeError("No raw data rows were written.")

    # ------------------------------------------------------------
    # "Summary" sheet (specific county/state)
    # ------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum.cell(row=1, column=1, value="Year")

    ws_sum.cell(row=1, column=2, value="County In (Exemptions)")
    ws_sum.cell(row=1, column=3, value="County Out (Exemptions)")
    ws_sum.cell(row=1, column=4, value="Net (In - Out) (Exemptions)")
    ws_sum.cell(row=1, column=5, value="Net per Day (Exemptions)")

    ws_sum.cell(row=1, column=6, value="County In (Returns)")
    ws_sum.cell(row=1, column=7, value="County Out (Returns)")
    ws_sum.cell(row=1, column=8, value="Net (In - Out) (Returns)")
    ws_sum.cell(row=1, column=9, value="Net per Day (Returns)")

    ws_sum.cell(row=1, column=10, value="County In (AGI)")
    ws_sum.cell(row=1, column=11, value="County Out (AGI)")
    ws_sum.cell(row=1, column=12, value="Net (In - Out) (AGI)")
    ws_sum.cell(row=1, column=13, value="Net per Day (AGI)")

    # Raw-data column indices (based on how ws_raw is constructed)
    COL_EXEMPTIONS = 9  # I
    COL_RETURNS = 8     # H
    COL_AGI = 10        # J

    TARGET_COUNTY_CODE = 119
    TARGET_STATE_CODE = 37

    out_row = 2
    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy
        dpy = days_in_year(year_value)

        county_in_ex = 0.0
        county_out_ex = 0.0
        county_in_ret = 0.0
        county_out_ret = 0.0
        county_in_agi = 0.0
        county_out_agi = 0.0

        for r in range(2, raw_last_row + 1):
            if ws_raw.cell(row=r, column=1).value != year_value:
                continue

            raw_col_E = ws_raw.cell(row=r, column=5).value  # to county
            raw_col_D = ws_raw.cell(row=r, column=4).value  # to state
            raw_col_C = ws_raw.cell(row=r, column=3).value  # from county
            raw_col_B = ws_raw.cell(row=r, column=2).value  # from state

            code_E = to_int_like(raw_col_E)
            code_D = to_int_like(raw_col_D)
            code_C = to_int_like(raw_col_C)
            code_B = to_int_like(raw_col_B)

            in_match = (code_E == TARGET_COUNTY_CODE and code_D == TARGET_STATE_CODE)
            out_match = (code_C == TARGET_COUNTY_CODE and code_B == TARGET_STATE_CODE)

            if in_match:
                county_in_ex += to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value)
                county_in_ret += to_float(ws_raw.cell(row=r, column=COL_RETURNS).value)
                county_in_agi += to_float(ws_raw.cell(row=r, column=COL_AGI).value)

            if out_match:
                county_out_ex += to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value)
                county_out_ret += to_float(ws_raw.cell(row=r, column=COL_RETURNS).value)
                county_out_agi += to_float(ws_raw.cell(row=r, column=COL_AGI).value)

        net_ex = county_in_ex - county_out_ex
        net_ret = county_in_ret - county_out_ret
        net_agi = county_in_agi - county_out_agi

        ws_sum.cell(row=out_row, column=1, value=year_value)

        ws_sum.cell(row=out_row, column=2, value=county_in_ex)
        ws_sum.cell(row=out_row, column=3, value=county_out_ex)
        ws_sum.cell(row=out_row, column=4, value=net_ex)
        ws_sum.cell(row=out_row, column=5, value=net_ex / dpy)

        ws_sum.cell(row=out_row, column=6, value=county_in_ret)
        ws_sum.cell(row=out_row, column=7, value=county_out_ret)
        ws_sum.cell(row=out_row, column=8, value=net_ret)
        ws_sum.cell(row=out_row, column=9, value=net_ret / dpy)

        ws_sum.cell(row=out_row, column=10, value=county_in_agi)
        ws_sum.cell(row=out_row, column=11, value=county_out_agi)
        ws_sum.cell(row=out_row, column=12, value=net_agi)
        ws_sum.cell(row=out_row, column=13, value=net_agi / dpy)

        out_row += 1

    # ------------------------------------------------------------
    # "Sum_total" sheet
    # Instead of SUMIF, use MAXIF-like logic:
    # For each year, among rows matching the "US+Foreign total" text
    # and the county/state direction, take the MAX of values rather than SUM.
    # ------------------------------------------------------------
    ws_sum_total = wb.create_sheet("Sum_total")
    ws_sum_total.cell(row=1, column=1, value="Year")

    ws_sum_total.cell(row=1, column=2, value="County In (Returns, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=3, value="County Out (Returns, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=4, value="Net (In - Out) (Returns)")
    ws_sum_total.cell(row=1, column=5, value="Net per Day (Returns)")

    ws_sum_total.cell(row=1, column=6, value="County In (Exemptions, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=7, value="County Out (Exemptions, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=8, value="Net (In - Out) (Exemptions)")
    ws_sum_total.cell(row=1, column=9, value="Net per Day (Exemptions)")

    ws_sum_total.cell(row=1, column=10, value="County In (AGI, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=11, value="County Out (AGI, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=12, value="Net (In - Out) (AGI)")
    ws_sum_total.cell(row=1, column=13, value="Net per Day (AGI)")

    # Update this if needed based on your raw header row.
    COL_COUNTY_NAME = 7
    REQUIRED_WORDS = ["us", "total", "foreign"]  # all must appear (case-insensitive)

    out_row = 2
    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy
        dpy = days_in_year(year_value)

        # MAXIF semantics: initialize to very small numbers so any real value replaces them
        in_ret = float("-inf")
        out_ret = float("-inf")
        in_ex = float("-inf")
        out_ex = float("-inf")
        in_agi = float("-inf")
        out_agi = float("-inf")

        for r in range(2, raw_last_row + 1):
            if ws_raw.cell(row=r, column=1).value != year_value:
                continue

            county_name_val = ws_raw.cell(row=r, column=COL_COUNTY_NAME).value
            if county_name_val is None:
                continue

            cn = str(county_name_val).lower()
            if not all(word in cn for word in REQUIRED_WORDS):
                continue

            raw_col_E = ws_raw.cell(row=r, column=5).value  # to county
            raw_col_D = ws_raw.cell(row=r, column=4).value  # to state
            raw_col_C = ws_raw.cell(row=r, column=3).value  # from county
            raw_col_B = ws_raw.cell(row=r, column=2).value  # from state

            code_E = to_int_like(raw_col_E)
            code_D = to_int_like(raw_col_D)
            code_C = to_int_like(raw_col_C)
            code_B = to_int_like(raw_col_B)

            in_match = (code_E == TARGET_COUNTY_CODE and code_D == TARGET_STATE_CODE)
            out_match = (code_C == TARGET_COUNTY_CODE and code_B == TARGET_STATE_CODE)

            if in_match:
                in_ret = max(in_ret, to_float(ws_raw.cell(row=r, column=COL_RETURNS).value))
                in_ex = max(in_ex, to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value))
                in_agi = max(in_agi, to_float(ws_raw.cell(row=r, column=COL_AGI).value))

            if out_match:
                out_ret = max(out_ret, to_float(ws_raw.cell(row=r, column=COL_RETURNS).value))
                out_ex = max(out_ex, to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value))
                out_agi = max(out_agi, to_float(ws_raw.cell(row=r, column=COL_AGI).value))

        # If nothing matched for a year, convert -inf to 0.0
        if in_ret == float("-inf"): in_ret = 0.0
        if out_ret == float("-inf"): out_ret = 0.0
        if in_ex == float("-inf"): in_ex = 0.0
        if out_ex == float("-inf"): out_ex = 0.0
        if in_agi == float("-inf"): in_agi = 0.0
        if out_agi == float("-inf"): out_agi = 0.0

        net_ret = in_ret - out_ret
        net_ex = in_ex - out_ex
        net_agi = in_agi - out_agi

        ws_sum_total.cell(row=out_row, column=1, value=year_value)

        ws_sum_total.cell(row=out_row, column=2, value=in_ret)
        ws_sum_total.cell(row=out_row, column=3, value=out_ret)
        ws_sum_total.cell(row=out_row, column=4, value=net_ret)
        ws_sum_total.cell(row=out_row, column=5, value=net_ret / dpy)

        ws_sum_total.cell(row=out_row, column=6, value=in_ex)
        ws_sum_total.cell(row=out_row, column=7, value=out_ex)
        ws_sum_total.cell(row=out_row, column=8, value=net_ex)
        ws_sum_total.cell(row=out_row, column=9, value=net_ex / dpy)

        ws_sum_total.cell(row=out_row, column=10, value=in_agi)
        ws_sum_total.cell(row=out_row, column=11, value=out_agi)
        ws_sum_total.cell(row=out_row, column=12, value=net_agi)
        ws_sum_total.cell(row=out_row, column=13, value=net_agi / dpy)

        out_row += 1

    wb.save(out_xlsx)
    print("Wrote:", out_xlsx)
