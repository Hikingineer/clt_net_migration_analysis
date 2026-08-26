import os
import requests
import pandas as pd
from openpyxl import Workbook

# ---- config ----
states_arg = "all"  # e.g. "all" or ["nc"] (string below supports "all" or a single abbr, or list)

overall_start_yy = 11
overall_end_yy = 22  # so last filename 2223nc.xlsx

IN_SHEET = "County Inflow"
OUT_SHEET = "County Outflow"

STATE_IN_SHEET = "State Inflow"
STATE_OUT_SHEET = "State Outflow"

URL_BASE = "https://www.irs.gov/pub/irs-soi/"
DOWNLOAD_DIR = "downloads_irs_soi"
os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def download_excel(url, out_path):
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with open(out_path, "wb") as f:
        f.write(r.content)


def ext_for_startyy(startyy: int) -> str:
    # xls till 19 (startyy <= 19); xlsx from 20 (startyy >= 20)
    return ".xlsx" if startyy >= 20 else ".xls"


def url_for_year_state(startyy: int, state_code: str):
    endyy = startyy + 1
    ext = ext_for_startyy(startyy)
    filename = f"{startyy}{endyy}{state_code}{ext}"
    return URL_BASE + filename, filename


def read_sheet(path, sheet_name):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return pd.read_excel(path, sheet_name=sheet_name, engine="xlrd")
    return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")


def to_float(x):
    try:
        if x is None:
            return 0.0
        if isinstance(x, str) and x.strip() == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0


def to_int_like(x):
    try:
        if x is None:
            return None
        return int(float(x))
    except Exception:
        return None


def days_in_year(y: int) -> int:
    return 366 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 365


def get_states(states_arg):
    all_states = [
        "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
        "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH",
        "NJ","NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX",
        "UT","VT","VA","WA","WV","WI","WY"
    ]

    if states_arg is None:
        return [s.lower() for s in all_states]

    if isinstance(states_arg, str):
        s = states_arg.strip().lower()
        if s == "all":
            return [x.lower() for x in all_states]
        return [s]

    cleaned = [s.strip().lower() for s in states_arg]
    invalid = [s for s in cleaned if s.upper() not in all_states]
    if invalid:
        raise ValueError(f"Invalid state abbreviation(s): {invalid}")
    return cleaned


def get_col_idx_by_header(ws, header_row, header_text):
    header_text = str(header_text).strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        if str(v).strip().lower() == header_text:
            return c
    raise KeyError(f"Header not found: {header_text}")


STATE_ABBR_TO_CODE = {
    "AL": 1, "AK": 2, "AZ": 4, "AR": 5, "CA": 6, "CO": 8, "CT": 9, "DE": 10, "FL": 11, "GA": 12,
    "HI": 13, "ID": 14, "IL": 15, "IN": 16, "IA": 17, "KS": 18, "KY": 19, "LA": 20, "ME": 21,
    "MD": 22, "MA": 23, "MI": 24, "MN": 25, "MS": 26, "MO": 27, "MT": 28, "NE": 31, "NV": 32,
    "NH": 33, "NJ": 34, "NM": 35, "NY": 36, "NC": 37, "ND": 38, "OH": 39, "OK": 40, "OR": 41,
    "PA": 42, "RI": 44, "SC": 45, "SD": 46, "TN": 47, "TX": 48, "UT": 49, "VT": 50, "VA": 51,
    "WA": 53, "WV": 54, "WI": 55, "WY": 56
}

states = get_states(states_arg)

for state in states:
    state = state.lower()
    out_xlsx = f"{state}_county_inflow_outflow.xlsx"

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "raw data"

    # ----------------------------
    # Build "raw data" (county-level)
    # ----------------------------
    wrote_headers = False
    next_row = 2  # leave row 1 for headers

    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy

        url, filename = url_for_year_state(startyy, state)
        local_path = os.path.join(DOWNLOAD_DIR, filename)

        print("Downloading:", url)
        download_excel(url, local_path)

        inflow = read_sheet(local_path, IN_SHEET)
        outflow = read_sheet(local_path, OUT_SHEET)

        if not wrote_headers:
            ws_raw.cell(row=1, column=1, value="Year")
            for j, colname in enumerate(inflow.columns.tolist(), start=2):
                ws_raw.cell(row=1, column=j, value=colname)
            wrote_headers = True
            next_row = 2

        # inflow rows
        for _, r in inflow.iterrows():
            ws_raw.cell(row=next_row, column=1, value=year_value)
            for j, val in enumerate(r.tolist(), start=2):
                ws_raw.cell(row=next_row, column=j, value=val)
            next_row += 1

        # outflow rows
        for _, r in outflow.iterrows():
            ws_raw.cell(row=next_row, column=1, value=year_value)
            for j, val in enumerate(r.tolist(), start=2):
                ws_raw.cell(row=next_row, column=j, value=val)
            next_row += 1

    raw_last_row = ws_raw.max_row
    if raw_last_row < 2:
        wb.save(out_xlsx)
        raise RuntimeError("No raw data rows were written.")

    # ----------------------------
    # Build "raw data (state-level)"
    # ----------------------------
    ws_state_raw = wb.create_sheet("raw data (state-level)")
    wrote_state_headers = False
    state_next_row = 2

    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy

        url, filename = url_for_year_state(startyy, state)
        local_path = os.path.join(DOWNLOAD_DIR, filename)

        # already downloaded above, but safe to re-read
        state_inflow = read_sheet(local_path, STATE_IN_SHEET)
        state_outflow = read_sheet(local_path, STATE_OUT_SHEET)

        if not wrote_state_headers:
            ws_state_raw.cell(row=1, column=1, value="Year")
            ws_state_raw.cell(row=1, column=2, value="Direction")
            for j, colname in enumerate(state_inflow.columns.tolist(), start=3):
                ws_state_raw.cell(row=1, column=j, value=colname)
            wrote_state_headers = True
            state_next_row = 2

        # "In" direction rows
        for _, rr in state_inflow.iterrows():
            ws_state_raw.cell(row=state_next_row, column=1, value=year_value)
            ws_state_raw.cell(row=state_next_row, column=2, value="In")
            for j, val in enumerate(rr.tolist(), start=3):
                ws_state_raw.cell(row=state_next_row, column=j, value=val)
            state_next_row += 1

        # "Out" direction rows
        for _, rr in state_outflow.iterrows():
            ws_state_raw.cell(row=state_next_row, column=1, value=year_value)
            ws_state_raw.cell(row=state_next_row, column=2, value="Out")
            for j, val in enumerate(rr.tolist(), start=3):
                ws_state_raw.cell(row=state_next_row, column=j, value=val)
            state_next_row += 1

    state_raw_last_row = ws_state_raw.max_row
    if state_raw_last_row < 2:
        wb.save(out_xlsx)
        raise RuntimeError("No state-level raw data rows were written.")

    # ------------------------------------------------------------
    # "Summary" sheet (original county-specific Mecklenburg summary)
    # ------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum.cell(row=1, column=1, value="Year")

    ws_sum.cell(row=1, column=2, value="County In (Exemptions)")
    ws_sum.cell(row=1, column=3, value="County Out (Exemptions)")
    ws_sum.cell(row=1, column=4, value="Net (In - Out) (Exemptions)")
    ws_sum.cell(row=1, column=5, value="Net per Day (Exemptions)")

    ws_sum.cell(row=1, column=6, value="County In (Returns)")
    ws_sum.cell(row=1, column=7, value="County Out (Returns)")
    ws_sum.cell(row=1, column=8, value="Net (In - Out) (Returns)")
    ws_sum.cell(row=1, column=9, value="Net per Day (Returns)")

    ws_sum.cell(row=1, column=10, value="County In (AGI)")
    ws_sum.cell(row=1, column=11, value="County Out (AGI)")
    ws_sum.cell(row=1, column=12, value="Net (In - Out) (AGI)")
    ws_sum.cell(row=1, column=13, value="Net per Day (AGI)")

    # Raw-data column indices (based on how ws_raw is constructed)
    # ws_raw columns:
    # 1: Year
    # 2..: from_state, from_county, to_state, to_county, then metrics...
    COL_EXEMPTIONS = 9  # I
    COL_RETURNS = 8     # H
    COL_AGI = 10        # J

    TARGET_COUNTY_CODE = 119
    TARGET_STATE_CODE = 37  # NC's state code in your request

    out_row = 2
    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy
        dpy = days_in_year(year_value)

        county_in_ex = 0.0
        county_out_ex = 0.0
        county_in_ret = 0.0
        county_out_ret = 0.0
        county_in_agi = 0.0
        county_out_agi = 0.0

        for r in range(2, raw_last_row + 1):
            if ws_raw.cell(row=r, column=1).value != year_value:
                continue

            raw_col_E = ws_raw.cell(row=r, column=5).value  # to county
            raw_col_D = ws_raw.cell(row=r, column=4).value  # to state
            raw_col_C = ws_raw.cell(row=r, column=3).value  # from county
            raw_col_B = ws_raw.cell(row=r, column=2).value  # from state

            code_E = to_int_like(raw_col_E)
            code_D = to_int_like(raw_col_D)
            code_C = to_int_like(raw_col_C)
            code_B = to_int_like(raw_col_B)

            in_match = (code_E == TARGET_COUNTY_CODE and code_D == TARGET_STATE_CODE)
            out_match = (code_C == TARGET_COUNTY_CODE and code_B == TARGET_STATE_CODE)

            if in_match:
                county_in_ex += to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value)
                county_in_ret += to_float(ws_raw.cell(row=r, column=COL_RETURNS).value)
                county_in_agi += to_float(ws_raw.cell(row=r, column=COL_AGI).value)

            if out_match:
                county_out_ex += to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value)
                county_out_ret += to_float(ws_raw.cell(row=r, column=COL_RETURNS).value)
                county_out_agi += to_float(ws_raw.cell(row=r, column=COL_AGI).value)

        net_ex = county_in_ex - county_out_ex
        net_ret = county_in_ret - county_out_ret
        net_agi = county_in_agi - county_out_agi

        ws_sum.cell(row=out_row, column=1, value=year_value)

        ws_sum.cell(row=out_row, column=2, value=county_in_ex)
        ws_sum.cell(row=out_row, column=3, value=county_out_ex)
        ws_sum.cell(row=out_row, column=4, value=net_ex)
        ws_sum.cell(row=out_row, column=5, value=net_ex / dpy)

        ws_sum.cell(row=out_row, column=6, value=county_in_ret)
        ws_sum.cell(row=out_row, column=7, value=county_out_ret)
        ws_sum.cell(row=out_row, column=8, value=net_ret)
        ws_sum.cell(row=out_row, column=9, value=net_ret / dpy)

        ws_sum.cell(row=out_row, column=10, value=county_in_agi)
        ws_sum.cell(row=out_row, column=11, value=county_out_agi)
        ws_sum.cell(row=out_row, column=12, value=net_agi)
        ws_sum.cell(row=out_row, column=13, value=net_agi / dpy)

        out_row += 1

    # ------------------------------------------------------------
    # "Sum_total" sheet (your original county total logic)
    # ------------------------------------------------------------
    ws_sum_total = wb.create_sheet("Sum_total")
    ws_sum_total.cell(row=1, column=1, value="Year")

    ws_sum_total.cell(row=1, column=2, value="County In (Returns, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=3, value="County Out (Returns, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=4, value="Net (In - Out) (Returns)")
    ws_sum_total.cell(row=1, column=5, value="Net per Day (Returns)")

    ws_sum_total.cell(row=1, column=6, value="County In (Exemptions, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=7, value="County Out (Exemptions, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=8, value="Net (In - Out) (Exemptions)")
    ws_sum_total.cell(row=1, column=9, value="Net per Day (Exemptions)")

    ws_sum_total.cell(row=1, column=10, value="County In (AGI, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=11, value="County Out (AGI, US+Foreign total row)")
    ws_sum_total.cell(row=1, column=12, value="Net (In - Out) (AGI)")
    ws_sum_total.cell(row=1, column=13, value="Net per Day (AGI)")

    COL_COUNTY_NAME = 7
    REQUIRED_WORDS = ["us", "total", "foreign"]

    out_row = 2
    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy
        dpy = days_in_year(year_value)

        in_ret = float("-inf")
        out_ret = float("-inf")
        in_ex = float("-inf")
        out_ex = float("-inf")
        in_agi = float("-inf")
        out_agi = float("-inf")

        for r in range(2, raw_last_row + 1):
            if ws_raw.cell(row=r, column=1).value != year_value:
                continue

            county_name_val = ws_raw.cell(row=r, column=COL_COUNTY_NAME).value
            if county_name_val is None:
                continue

            cn = str(county_name_val).lower()
            if not all(word in cn for word in REQUIRED_WORDS):
                continue

            raw_col_E = ws_raw.cell(row=r, column=5).value  # to county
            raw_col_D = ws_raw.cell(row=r, column=4).value  # to state
            raw_col_C = ws_raw.cell(row=r, column=3).value  # from county
            raw_col_B = ws_raw.cell(row=r, column=2).value  # from state

            code_E = to_int_like(raw_col_E)
            code_D = to_int_like(raw_col_D)
            code_C = to_int_like(raw_col_C)
            code_B = to_int_like(raw_col_B)

            in_match = (code_E == TARGET_COUNTY_CODE and code_D == TARGET_STATE_CODE)
            out_match = (code_C == TARGET_COUNTY_CODE and code_B == TARGET_STATE_CODE)

            if in_match:
                in_ret = max(in_ret, to_float(ws_raw.cell(row=r, column=COL_RETURNS).value))
                in_ex = max(in_ex, to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value))
                in_agi = max(in_agi, to_float(ws_raw.cell(row=r, column=COL_AGI).value))

            if out_match:
                out_ret = max(out_ret, to_float(ws_raw.cell(row=r, column=COL_RETURNS).value))
                out_ex = max(out_ex, to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value))
                out_agi = max(out_agi, to_float(ws_raw.cell(row=r, column=COL_AGI).value))

        if in_ret == float("-inf"): in_ret = 0.0
        if out_ret == float("-inf"): out_ret = 0.0
        if in_ex == float("-inf"): in_ex = 0.0
        if out_ex == float("-inf"): out_ex = 0.0
        if in_agi == float("-inf"): in_agi = 0.0
        if out_agi == float("-inf"): out_agi = 0.0

        net_ret = in_ret - out_ret
        net_ex = in_ex - out_ex
        net_agi = in_agi - out_agi

        ws_sum_total.cell(row=out_row, column=1, value=year_value)

        ws_sum_total.cell(row=out_row, column=2, value=in_ret)
        ws_sum_total.cell(row=out_row, column=3, value=out_ret)
        ws_sum_total.cell(row=out_row, column=4, value=net_ret)
        ws_sum_total.cell(row=out_row, column=5, value=net_ret / dpy)

        ws_sum_total.cell(row=out_row, column=6, value=in_ex)
        ws_sum_total.cell(row=out_row, column=7, value=out_ex)
        ws_sum_total.cell(row=out_row, column=8, value=net_ex)
        ws_sum_total.cell(row=out_row, column=9, value=net_ex / dpy)

        ws_sum_total.cell(row=out_row, column=10, value=in_agi)
        ws_sum_total.cell(row=out_row, column=11, value=out_agi)
        ws_sum_total.cell(row=out_row, column=12, value=net_agi)
        ws_sum_total.cell(row=out_row, column=13, value=net_agi / dpy)

        out_row += 1

    # ------------------------------------------------------------
    # NEW: State-level compilation + summary + sense check
    # ------------------------------------------------------------
    ws_state_sum = wb.create_sheet("Sum_state_level")
    ws_state_sum.cell(row=1, column=1, value="Year")
    ws_state_sum.cell(row=1, column=2, value="State In (Exemptions)")
    ws_state_sum.cell(row=1, column=3, value="State Out (Exemptions)")
    ws_state_sum.cell(row=1, column=4, value="Net (In - Out) (Exemptions)")
    ws_state_sum.cell(row=1, column=5, value="Net per Day (Exemptions)")

    ws_state_sum.cell(row=1, column=6, value="State In (Returns)")
    ws_state_sum.cell(row=1, column=7, value="State Out (Returns)")
    ws_state_sum.cell(row=1, column=8, value="Net (In - Out) (Returns)")
    ws_state_sum.cell(row=1, column=9, value="Net per Day (Returns)")

    ws_state_sum.cell(row=1, column=10, value="State In (AGI)")
    ws_state_sum.cell(row=1, column=11, value="State Out (AGI)")
    ws_state_sum.cell(row=1, column=12, value="Net (In - Out) (AGI)")
    ws_state_sum.cell(row=1, column=13, value="Net per Day (AGI)")

    # locate columns in state raw by header text (safer than hard-coded)
    # NOTE: headers in your state-level sheets might be e.g. "Exemptions", "Returns", "AGI".
    # If yours differ, update these header strings.
    COL_S_EX = get_col_idx_by_header(ws_state_raw, 1, "Exemptions")
    COL_S_RET = get_col_idx_by_header(ws_state_raw, 1, "Returns")
    COL_S_AGI = get_col_idx_by_header(ws_state_raw, 1, "AGI")

    target_state_code_for_sense = STATE_ABBR_TO_CODE[state.upper()]

    out_row = 2
    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy
        dpy = days_in_year(year_value)

        state_in_ex = 0.0
        state_out_ex = 0.0
        state_in_ret = 0.0
        state_out_ret = 0.0
        state_in_agi = 0.0
        state_out_agi = 0.0

        for r in range(2, state_raw_last_row + 1):
            if ws_state_raw.cell(row=r, column=1).value != year_value:
                continue

            direction = ws_state_raw.cell(row=r, column=2).value
            if direction is None:
                continue
            direction = str(direction).strip().lower()

            if direction == "in":
                state_in_ex += to_float(ws_state_raw.cell(row=r, column=COL_S_EX).value)
                state_in_ret += to_float(ws_state_raw.cell(row=r, column=COL_S_RET).value)
                state_in_agi += to_float(ws_state_raw.cell(row=r, column=COL_S_AGI).value)
            elif direction == "out":
                state_out_ex += to_float(ws_state_raw.cell(row=r, column=COL_S_EX).value)
                state_out_ret += to_float(ws_state_raw.cell(row=r, column=COL_S_RET).value)
                state_out_agi += to_float(ws_state_raw.cell(row=r, column=COL_S_AGI).value)

        net_ex = state_in_ex - state_out_ex
        net_ret = state_in_ret - state_out_ret
        net_agi = state_in_agi - state_out_agi

        ws_state_sum.cell(row=out_row, column=1, value=year_value)

        ws_state_sum.cell(row=out_row, column=2, value=state_in_ex)
        ws_state_sum.cell(row=out_row, column=3, value=state_out_ex)
        ws_state_sum.cell(row=out_row, column=4, value=net_ex)
        ws_state_sum.cell(row=out_row, column=5, value=net_ex / dpy)

        ws_state_sum.cell(row=out_row, column=6, value=state_in_ret)
        ws_state_sum.cell(row=out_row, column=7, value=state_out_ret)
        ws_state_sum.cell(row=out_row, column=8, value=net_ret)
        ws_state_sum.cell(row=out_row, column=9, value=net_ret / dpy)

        ws_state_sum.cell(row=out_row, column=10, value=state_in_agi)
        ws_state_sum.cell(row=out_row, column=11, value=state_out_agi)
        ws_state_sum.cell(row=out_row, column=12, value=net_agi)
        ws_state_sum.cell(row=out_row, column=13, value=net_agi / dpy)

        out_row += 1

    # Sense check: compare state net derived from state-level vs derived from county-level
    ws_check = wb.create_sheet("SenseCheck_state")
    ws_check.cell(row=1, column=1, value="Year")
    ws_check.cell(row=1, column=2, value="StateNet_from_state-level_Ex")
    ws_check.cell(row=1, column=3, value="StateNet_from_county-level_Ex")
    ws_check.cell(row=1, column=4, value="Diff_Ex")

    ws_check.cell(row=1, column=5, value="StateNet_from_state-level_Ret")
    ws_check.cell(row=1, column=6, value="StateNet_from_county-level_Ret")
    ws_check.cell(row=1, column=7, value="Diff_Ret")

    ws_check.cell(row=1, column=8, value="StateNet_from_state-level_AGI")
    ws_check.cell(row=1, column=9, value="StateNet_from_county-level_AGI")
    ws_check.cell(row=1, column=10, value="Diff_AGI")

    out_row = 2
    for startyy in range(overall_start_yy, overall_end_yy + 1):
        endyy = startyy + 1
        year_value = 2000 + endyy
        dpy = days_in_year(year_value)

        # county-inferred state net
        county_state_in_ex = 0.0
        county_state_out_ex = 0.0
        county_state_in_ret = 0.0
        county_state_out_ret = 0.0
        county_state_in_agi = 0.0
        county_state_out_agi = 0.0

        for r in range(2, raw_last_row + 1):
            if ws_raw.cell(row=r, column=1).value != year_value:
                continue

            raw_col_D = ws_raw.cell(row=r, column=4).value  # to state
            raw_col_B = ws_raw.cell(row=r, column=2).value  # from state

            code_D = to_int_like(raw_col_D)
            code_B = to_int_like(raw_col_B)

            if code_D == target_state_code_for_sense:
                county_state_in_ex += to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value)
                county_state_in_ret += to_float(ws_raw.cell(row=r, column=COL_RETURNS).value)
                county_state_in_agi += to_float(ws_raw.cell(row=r, column=COL_AGI).value)

            if code_B == target_state_code_for_sense:
                county_state_out_ex += to_float(ws_raw.cell(row=r, column=COL_EXEMPTIONS).value)
                county_state_out_ret += to_float(ws_raw.cell(row=r, column=COL_RETURNS).value)
                county_state_out_agi += to_float(ws_raw.cell(row=r, column=COL_AGI).value)

        county_net_ex = county_state_in_ex - county_state_out_ex
        county_net_ret = county_state_in_ret - county_state_out_ret
        county_net_agi = county_state_in_agi - county_state_out_agi

        # state-level net pulled from ws_state_sum
        state_level_row = 2 + (startyy - overall_start_yy)
        state_net_ex = to_float(ws_state_sum.cell(row=state_level_row, column=4).value)
        state_net_ret = to_float(ws_state_sum.cell(row=state_level_row, column=8).value)
        state_net_agi = to_float(ws_state_sum.cell(row=state_level_row, column=12).value)

        ws_check.cell(row=out_row, column=1, value=year_value)
        ws_check.cell(row=out_row, column=2, value=state_net_ex)
        ws_check.cell(row=out_row, column=3, value=county_net_ex)
        ws_check.cell(row=out_row, column=4, value=state_net_ex - county_net_ex)

        ws_check.cell(row=out_row, column=5, value=state_net_ret)
        ws_check.cell(row=out_row, column=6, value=county_net_ret)
        ws_check.cell(row=out_row, column=7, value=state_net_ret - county_net_ret)

        ws_check.cell(row=out_row, column=8, value=state_net_agi)
        ws_check.cell(row=out_row, column=9, value=county_net_agi)
        ws_check.cell(row=out_row, column=10, value=state_net_agi - county_net_agi)

        out_row += 1

    wb.save(out_xlsx)
    print("Wrote:", out_xlsx)
